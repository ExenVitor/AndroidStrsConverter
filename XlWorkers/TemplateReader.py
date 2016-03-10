# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from XlWorkers.Config import Config
import XlWorkers.Utils as Utils

__author__ = 'Vitor Chen'


class TemplateReader(object):
    def __init__(self, filename):
        super().__init__()
        self._lang_dict = {}

        wb = load_workbook(filename, read_only=True)
        ws = wb.active
        for content_row in ws.iter_rows(row_offset=Config.CONTENT_ROW_START - 1):
            key_cell = content_row[Config.KEY_COL_POS - 1]
            res_key = key_cell.value
            if Utils.is_str_valid(res_key):
                for cell_i in range(Config.LANG_COL_START - 1, len(content_row)):
                    cell = content_row[cell_i]
                    lang_col_info = Config.LANG_COL_POS_MAP.get(cell.column)
                    if lang_col_info is not None:
                        self._set_lang_dict_value(lang_col_info.code, res_key, cell.value)

    def _set_lang_dict_value(self, lang_code, res_key, value):
        if self._lang_dict.get(lang_code) is None:
            self._lang_dict[lang_code] = {}

        res_str_dict = self._lang_dict.get(lang_code)
        res_str_dict[res_key] = value

    def get_res_str(self, lang_code, res_key):
        res_str_dict = self._lang_dict.get(lang_code)
        if res_str_dict is not None:
            return res_str_dict.get(res_key)
        return None

    def __str__(self, *args, **kwargs):
        return str(self._lang_dict)


if __name__ == '__main__':
    template = TemplateReader('../template.xlsx')
    print(template)
