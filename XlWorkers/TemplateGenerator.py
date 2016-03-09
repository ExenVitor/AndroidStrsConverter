# -*- coding: utf-8 -*-

from openpyxl import Workbook

from XlWorkers.Config import *
from XlWorkers.Entities import TransEntity

__author__ = 'Vitor Chen'


class TemplateGenerator(object):
    def __init__(self):
        super().__init__()
        self._column_titles = []
        self._trans_entities = []
        for col_info in Config.HEADER_COL_LIST:
            self._column_titles.append(col_info.display_name)

    def append_trans_entities(self, entities):
        if self.__verify_input_array(entities):
            for entity in entities:
                if isinstance(entity, TransEntity):
                    self._trans_entities.append(entity)

    def gen_template(self, output_path=None, lang_code=LangCode.EN.code):
        wb = Workbook()
        ws = wb.active
        ws.append(self._column_titles)

        self._write_res_trans_entities(ws, lang_code)

        if output_path is None:
            output_path = 'template.xlsx'
        wb.save(output_path)

    def _write_res_trans_entities(self, ws, lang_code):
        start_row = 2
        lang_cell_pos = Config.LANG_COL_CODE_MAP.get(lang_code)
        lang_col_pos = lang_cell_pos.col

        entities_count = len(self._trans_entities)

        if entities_count > 0:
            for i in range(entities_count):
                trans_entity = self._trans_entities[i]
                current_row = start_row + i

                key_cell = ws.cell(row=current_row, column=Config.KEY_COL_POS)
                key_cell.value = trans_entity.key

                trans_cell = ws.cell(row=current_row, column=lang_col_pos)
                trans_cell.value = trans_entity.trans_str

                if trans_entity.desc is not None:
                    desc_cell = ws.cell(row=current_row, column=Config.DESC_COL_POS)
                    desc_cell.value = trans_entity.desc

    @staticmethod
    def __verify_input_array(input_array):
        if isinstance(input_array, (list, tuple)):
            return True
        else:
            raise TypeError('Value must be a list, tuple. Supplied value is {0}'.format(
                type(input_array)))
